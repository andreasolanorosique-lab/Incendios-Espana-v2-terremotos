import os
import csv
import json
import re
import requests
import time
import xml.etree.ElementTree as ET

from math import (
    radians,
    sin,
    cos,
    sqrt,
    atan2,
    degrees,
    asin,
)

from openpyxl import load_workbook


# =====================================================
# CONFIGURACIÓN
# =====================================================

MAP_KEY = os.environ["FIRMS_MAP_KEY"]

SOURCES = [
    "VIIRS_SNPP_NRT",
    "VIIRS_NOAA20_NRT",
    "VIIRS_NOAA21_NRT",
]

BBOX = "-10,35,5,44"

DISTANCIA_AGRUPACION = 300      # metros



EFFIS_URL = (
    "https://services-eu1.arcgis.com/"
    "VC42ANIVJ5dUfvUn/ArcGIS/rest/services/"
    "Burned_Areas_EFFIS/FeatureServer/23/query"
)

print("Descargando datos NASA FIRMS...")

with open("fires.csv", "wb") as salida:

    primera = True

    for source in SOURCES:

        url = (
            "https://firms.modaps.eosdis.nasa.gov/api/area/csv/"
            f"{MAP_KEY}/{source}/{BBOX}/1"
        )

        print(f"  -> {source}")
    
        respuesta = None

        for intento in range(5):
            try:
                print(f"Intento {intento + 1}/5")

                respuesta = requests.get(url, timeout=120)
                respuesta.raise_for_status()
                break

            except requests.exceptions.RequestException as e:
                print(f"Error de conexión: {e}")

                if intento == 4:
                    raise

                print("Reintentando en 15 segundos...")
                time.sleep(15)

        print("Código:", respuesta.status_code)
        print("URL:", url)
        print("Primeros 300 caracteres:")
        print(repr(respuesta.text[:300]))

        lineas = respuesta.text.strip().splitlines()
        print("Número de líneas:", len(lineas))

        if len(lineas) <= 1:
            continue

        if primera:
            salida.write((lineas[0] + "\n").encode("utf-8"))
            primera = False

        for linea in lineas[1:]:
            salida.write((linea + "\n").encode("utf-8"))

print("Datos descargados correctamente.")

with open("fires.csv", "r", encoding="utf-8") as f:
    print("=== Primeras líneas de fires.csv ===")
    for i in range(10):
        print(f.readline().rstrip())
# =====================================================
# CREAR KML
# =====================================================

kml = ET.Element(
    "kml",
    xmlns="http://www.opengis.net/kml/2.2",
)

documento = ET.SubElement(kml, "Document")

ET.SubElement(
    documento,
    "name"
).text = "Incendios activos España"


BASE_ICONOS = (
    "https://andreasolanorosique-lab.github.io/"
    "Incendios-Espana-v2/icons/"
)


ESTILOS = {

    "pink": (
        "fire_pink.png",
        1.0,
    ),

    "green": (
        "fire_green.png",
        1.0,
    ),

    "yellow": (
        "fire_yellow.png",
        1.0,
    ),

    "orange": (
        "fire_orange.png",
        1.0,
    ),

    "red": (
        "fire_red.png",
        1.0,
    ),
}


for nombre_estilo, (icono, escala) in ESTILOS.items():

    estilo = ET.SubElement(
        documento,
        "Style",
        id=nombre_estilo,
    )

    icon_style = ET.SubElement(
        estilo,
        "IconStyle",
    )

    ET.SubElement(
        icon_style,
        "scale",
    ).text = str(escala)

    icon = ET.SubElement(
        icon_style,
        "Icon",
    )

    ET.SubElement(
        icon,
        "href",
    ).text = BASE_ICONOS + icono

    label = ET.SubElement(
        estilo,
        "LabelStyle",
    )

    ET.SubElement(
        label,
        "scale",
    ).text = "0"
# Estilo del área amarilla
estilo = ET.SubElement(
    documento,
    "Style",
    id="area_amarilla",
)

line_style = ET.SubElement(
    estilo,
    "LineStyle",
)

ET.SubElement(
    line_style,
    "color",
).text = "ff00ffff"

ET.SubElement(
    line_style,
    "width",
).text = "2"

poly_style = ET.SubElement(
    estilo,
    "PolyStyle",
)

ET.SubElement(
    poly_style,
    "color",
).text = "5500ffff"

# =====================================================
# FUNCIONES GEOGRÁFICAS
# =====================================================


def distancia_metros(lat1, lon1, lat2, lon2):

    radio_tierra = 6371000

    lat1 = radians(lat1)
    lon1 = radians(lon1)

    lat2 = radians(lat2)
    lon2 = radians(lon2)

    dlat = lat2 - lat1
    dlon = lon2 - lon1

    a = (
        sin(dlat / 2) ** 2
        + cos(lat1)
        * cos(lat2)
        * sin(dlon / 2) ** 2
    )

    c = 2 * atan2(
        sqrt(a),
        sqrt(1 - a),
    )

    return radio_tierra * c
def calcular_radio_grupo(grupo):
    """
    Calcula el centro del grupo y el radio estimado.
    Devuelve:
        lat_centro, lon_centro, radio
    """

    lat_centro = sum(p["lat"] for p in grupo) / len(grupo)
    lon_centro = sum(p["lon"] for p in grupo) / len(grupo)

    radio = 0

    for punto in grupo:

        distancia = distancia_metros(
            lat_centro,
            lon_centro,
            punto["lat"],
            punto["lon"],
        )

        if distancia > radio:
            radio = distancia

    # Margen de seguridad
    radio += 100

    # Límites
    radio = max(150, radio)
    radio = min(2500, radio)

    return lat_centro, lon_centro, radio

def crear_circulo(lat, lon, radio_m, pasos=36):
    radio_tierra = 6371000

    puntos = []

    lat1 = radians(lat)
    lon1 = radians(lon)

    for i in range(pasos + 1):
        angulo = radians(i * 360 / pasos)

        lat2 = asin(
            sin(lat1) * cos(radio_m / radio_tierra)
            + cos(lat1) * sin(radio_m / radio_tierra) * cos(angulo)
        )

        lon2 = lon1 + atan2(
            sin(angulo) * sin(radio_m / radio_tierra) * cos(lat1),
            cos(radio_m / radio_tierra) - sin(lat1) * sin(lat2),
        )

        puntos.append(
            (
                degrees(lat2),
                degrees(lon2),
            )
        )

    return puntos

def agrupar_focos(focos):

    grupos = []

    for foco in focos:

        encontrado = False

        for grupo in grupos:

            for referencia in grupo:

                distancia = distancia_metros(
                    foco["lat"],
                    foco["lon"],
                    referencia["lat"],
                    referencia["lon"],
                )

                if distancia <= DISTANCIA_AGRUPACION:

                    grupo.append(foco)
                    encontrado = True
                    break

            if encontrado:
                break

        if not encontrado:
            grupos.append([foco])

    return grupos


def cargar_municipios():

    print("Cargando municipios...")

    wb = load_workbook(
        "IGN_INFOGEO_MUNICIPIOS.xlsx",
        read_only=True,
        data_only=True,
    )

    ws = wb["IGN_INFOGEO_MUNICIPIOS"]

    municipios = []

    encabezados = [
        c.value
        for c in next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
            )
        )
    ]

    idx_nombre = encabezados.index("Nombre")
    idx_mapa = encabezados.index("Ver en mapa")

    for fila in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):

        nombre = fila[idx_nombre]
        url = fila[idx_mapa]

        if not nombre:
            continue

        if not url:
            continue

        coincidencia = re.search(
            r"center=([-\d\.]+),([-\d\.]+)",
            url,
        )

        if coincidencia is None:
            continue

        lon = float(coincidencia.group(1))
        lat = float(coincidencia.group(2))

        municipios.append(
            {
                "nombre": nombre,
                "lat": lat,
                "lon": lon,
            }
        )

    wb.close()

    print(f"Municipios cargados: {len(municipios)}")

    return municipios


def municipio_mas_cercano(lat, lon, municipios):

    mejor = None
    mejor_distancia = float("inf")

    for municipio in municipios:

        distancia = distancia_metros(
            lat,
            lon,
            municipio["lat"],
            municipio["lon"],
        )

        if distancia < mejor_distancia:

            mejor_distancia = distancia
            mejor = municipio

    return mejor, mejor_distancia
def buscar_effis(lat, lon):

    params = {
        "f": "json",
        "geometry": f"{lon},{lat}",
        "geometryType": "esriGeometryPoint",
        "spatialRel": "esriSpatialRelIntersects",
        "inSR": "4326",
        "outFields": "*",
        "returnGeometry": "true"
    }

    r = requests.get(
        EFFIS_URL,
        params=params,
        timeout=30
    )

    r.raise_for_status()

    datos = r.json()

    if "features" not in datos:
        return None

    if len(datos["features"]) == 0:
        return None

    return datos["features"][0]

# =====================================================
# LEER TODOS LOS FOCOS NASA
# =====================================================

focos = []

with open(
    "fires.csv",
    encoding="utf-8",
) as f:

    lector = csv.DictReader(f)

    for row in lector:

        lat = row.get("latitude")
        lon = row.get("longitude")

        if not lat or not lon:
            continue

        try:
            frp = float(row.get("frp") or 0)
        except ValueError:
            frp = 0

        focos.append(
            {
                "lat": float(lat),
                "lon": float(lon),
                "frp": frp,
                "row": row,
            }
        )

print(f"Focos leídos: {len(focos)}")
# =====================================================
# AGRUPAR FOCOS
# =====================================================

municipios = cargar_municipios()

grupos = []
grupos = agrupar_focos(focos)
print(f"Grupos detectados: {len(grupos)}")

# Elimina focos aislados muy débiles
#grupos = [
#    grupo
#    for grupo in grupos
#    if len(grupo) > 1
#    or max(f["frp"] for f in grupo) >= 50
#]

print(f"Grupos tras el filtrado: {len(grupos)}")

# =====================================================
# MENSAJE SI NO HAY INCENDIOS
# =====================================================

if len(grupos) == 0:
    print(">>> Entrando en el bloque SIN INCENDIOS <<<")
    
    overlay = ET.SubElement(documento, "ScreenOverlay")

    ET.SubElement(
        overlay,
        "name"
    ).text = "Sin incendios"

    icon = ET.SubElement(
        overlay,
        "Icon"
    )

    ET.SubElement(
        icon,
        "href"
    ).text = (
        "https://andreasolanorosique-lab.github.io/"
        "Incendios-Espana-v2/icons/no_incendios.png"
    )

    ET.SubElement(
        overlay,
        "overlayXY",
        x="0",
        y="0",
        xunits="fraction",
        yunits="fraction"
    )

    ET.SubElement(
        overlay,
        "screenXY",
        x="0.02",
        y="0.02",
        xunits="fraction",
        yunits="fraction"
    )

    ET.SubElement(
        overlay,
        "rotationXY",
        x="0",
        y="0",
        xunits="fraction",
        yunits="fraction"
    )

    ET.SubElement(
        overlay,
        "size",
        x="0",
        y="0.45",
        xunits="fraction",
        yunits="fraction"
    )
  
# =====================================================
# CREAR PLACEMARKS
# =====================================================

for grupo in grupos:
    print("Creando incendio...")
    cantidad = len(grupo)

    lat = sum(f["lat"] for f in grupo) / cantidad
    lon = sum(f["lon"] for f in grupo) / cantidad
    # Calcular el radio estimado del incendio
    radio = 0

    for foco in grupo:

        distancia = distancia_metros(
        lat,
        lon,
        foco["lat"],
        foco["lon"],
    )

        if distancia > radio:
            radio = distancia

    # Añadir margen de seguridad
    radio += 100

    # Limitar el tamaño del círculo
    radio = max(150, radio)
    radio = min(2500, radio)
    foco_principal = max(
        grupo,
        key=lambda f: f["frp"],
    )

    row = foco_principal["row"]
    frp = foco_principal["frp"]

    municipio, distancia = municipio_mas_cercano(
        lat,
        lon,
        municipios,
    )

    if distancia >= 1000:
        distancia_txt = f"{distancia / 1000:.1f} km"
    else:
        distancia_txt = f"{int(distancia)} m"


    # -------------------------------
    # Color del icono
    # -------------------------------

    if frp < 50:
        style = "#pink"
        confianza = "Baja"

    elif frp < 100:
        style = "#yellow"
        confianza = "Media"

    elif frp < 200:
        style = "#orange"
        confianza = "Alta"

    else:
        style = "#red"
        confianza = "Muy alta"
    # =====================================================
    # ÁREA ESTIMADA DEL INCENDIO
    # =====================================================

    placemark_area = ET.SubElement(
        documento,
        "Placemark",
    )

    ET.SubElement(
        placemark_area,
        "styleUrl",
    ).text = "#area_amarilla"

    polygon = ET.SubElement(
        placemark_area,
        "Polygon",
    )

    outer = ET.SubElement(
        polygon,
        "outerBoundaryIs",
    )

    ring = ET.SubElement(
        outer,
        "LinearRing",
    )

    coords = ET.SubElement(
        ring,
        "coordinates",
    )

    puntos = crear_circulo(
        lat,
        lon,
        radio,
    )

    coords.text = " ".join(
        f"{lon},{lat},0"
        for lat, lon in puntos
    )

    placemark = ET.SubElement(
        documento,
        "Placemark",
    )

    ET.SubElement(
        placemark,
        "styleUrl",
    ).text = style

    ET.SubElement(
        placemark,
        "name",
    ).text = ""
    descripcion = f"""
    <![CDATA[
    <h2>🔥 Incendio activo</h2>

    <table border="0" cellpadding="4">

        <tr>
            <td><b>Focos agrupados</b></td>
            <td>{cantidad}</td>
        </tr>

        <tr>
            <td><b>FRP máximo</b></td>
            <td>{frp:.1f} MW</td>
        </tr>

        <tr>
            <td><b>Confianza</b></td>
            <td>{confianza}</td>
        </tr>

        <tr>
            <td><b>Fecha</b></td>
            <td>{row.get("acq_date","")}</td>
        </tr>

        <tr>
            <td><b>Hora</b></td>
            <td>{row.get("acq_time","")} UTC</td>
        </tr>

        <tr>
            <td><b>Satélite</b></td>
            <td>{row.get("satellite","")}</td>
        </tr>

        <tr>
            <td><b>Instrumento</b></td>
            <td>{row.get("instrument","")}</td>
        </tr>

        <tr>
            <td><b>Población más cercana</b></td>
            <td>{municipio["nombre"]}</td>
        </tr>

        <tr>
            <td><b>Distancia</b></td>
            <td>{distancia_txt}</td>
        </tr>

        <tr>
            <td><b>Latitud</b></td>
            <td>{lat:.6f}</td>
        </tr>

        <tr>
            <td><b>Longitud</b></td>
            <td>{lon:.6f}</td>
        </tr>

    </table>

    ]]>
    """

    ET.SubElement(
        placemark,
        "description",
    ).text = descripcion

    punto = ET.SubElement(
        placemark,
        "Point",
    )

    ET.SubElement(
        punto,
        "coordinates",
    ).text = f"{lon},{lat},0"


# =====================================================
# CARGAR RED DE GASODUCTOS
# =====================================================

print("Cargando red de gasoductos...")

with open(
    "infraestructuras/gasoductos/gasoductos.json",
    encoding="utf-8",
) as f:

    gasoductos = json.load(f)


# =====================================================
# DIBUJAR RED DE GASODUCTOS
# =====================================================

carpeta_gas = ET.SubElement(
    documento,
    "Folder",
)

ET.SubElement(
    carpeta_gas,
    "name",
).text = "Gasoductos"


for tramo in gasoductos:

    coordenadas = tramo.get("coordenadas", [])

    if len(coordenadas) < 2:
        continue

    placemark = ET.SubElement(
        carpeta_gas,
        "Placemark",
    )

    ET.SubElement(
        placemark,
        "name",
    ).text = tramo.get(
        "nombre",
        "Gasoducto",
    )

    estilo = ET.SubElement(
        placemark,
        "Style",
    )

    linea = ET.SubElement(
        estilo,
        "LineStyle",
    )

    ET.SubElement(
        linea,
        "color",
    ).text = "ff0000ff"

    ET.SubElement(
        linea,
        "width",
    ).text = "4"

    linestring = ET.SubElement(
        placemark,
        "LineString",
    )

    ET.SubElement(
        linestring,
        "tessellate",
    ).text = "1"

    ET.SubElement(
        linestring,
        "altitudeMode",
    ).text = "clampToGround"

    ET.SubElement(
        linestring,
        "coordinates",
    ).text = "\n".join(
        f"{lon},{lat}"
        for lon, lat in coordenadas
    )

print(f"Gasoductos cargados: {len(gasoductos)}")

# =====================================================
# ESCRIBIR KML
# =====================================================

print("Generando archivo KML...")

tree = ET.ElementTree(kml)

try:
    ET.indent(
        tree,
        space="  ",
    )
except AttributeError:
    # Compatible con Python < 3.9
    pass

tree.write(
    "incendios_live.kml",
    encoding="utf-8",
    xml_declaration=True,
)

print("========================================")
print("KML generado correctamente")
print(f"Focos originales : {len(focos)}")
print(f"Grupos creados   : {len(grupos)}")
print(f"Gasoductos       : {len(gasoductos)}")
print("Archivo: incendios_live.kml")
print("========================================")
