
#!/usr/bin/env python3
"""
Genera catalogo_posiciones.json a partir del Excel
LISTADO POSICIONES(Orion).xlsx

Requisitos:
    pip install openpyxl
"""

from openpyxl import load_workbook
import json
from pathlib import Path

EXCEL = Path("LISTADO POSICIONES(Orion).xlsx")
SALIDA = Path("infraestructuras/posiciones/catalogo_posiciones.json")


def buscar_columna(cabeceras, texto):
    for i, h in enumerate(cabeceras):
        if h and texto.lower() in str(h).lower():
            return i
    raise ValueError(f"No se encontró la columna: {texto}")


def main():
    if not EXCEL.exists():
        raise FileNotFoundError(f"No existe {EXCEL}")

    wb = load_workbook(EXCEL, data_only=True)
    ws = wb[wb.sheetnames[0]]

    cabeceras = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    idx_codigo = buscar_columna(cabeceras, "posición")
    idx_desc = buscar_columna(cabeceras, "descrip")
    idx_tramo = buscar_columna(cabeceras, "tramo")
    idx_ct = buscar_columna(cabeceras, "centro")
    idx_x = buscar_columna(cabeceras, "coordenada x")
    idx_y = buscar_columna(cabeceras, "coordenada y")

    registros = []

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not any(fila):
            continue

        descripcion = fila[idx_desc]

        registros.append({
            "codigo": fila[idx_codigo],
            "tramo": fila[idx_tramo],
            "descripcion": descripcion,
            "poblacion": descripcion,
            "latitud": fila[idx_y],
            "longitud": fila[idx_x],
            "centro_transporte": fila[idx_ct]
        })

    SALIDA.parent.mkdir(parents=True, exist_ok=True)

    with SALIDA.open("w", encoding="utf-8") as f:
        json.dump(registros, f, ensure_ascii=False, indent=2)

    print(f"Generados {len(registros)} registros")
    print(f"Archivo creado: {SALIDA}")


if __name__ == "__main__":
    main()
